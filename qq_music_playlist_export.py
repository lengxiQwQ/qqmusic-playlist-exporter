# --- QQ音乐歌单导出工具 ---
# -*- coding: utf-8 -*-

"""
依赖（自行安装）：
    pip install requests
    pip install openpyxl
"""

__author__ = "lengxiQwQ"

import os
import re
import sys
import json
import csv
import platform
import subprocess
import requests

# --- 辅助函数 ---

# 去掉可能的 JSONP 包装，返回 JSON 字符串
def strip_jsonp(text):
    m = re.search(r'^[^\(]*\((\{.*\})\)\s*;?\s*$', text, flags=re.S)
    if m:
        return m.group(1)
    return text

# 将文件名中 Windows 不允许的字符替换为空格：< > : " / \ | ? * 以及控制字符（0-31），去掉首尾空白
def sanitize_filename(name):
    if not name:
        return ""
    name = re.sub(r'[\x00-\x1f]', ' ', name)
    name = re.sub(r'[<>:"/\\|?*]', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name or "playlist"

# 从用户输入中提取歌单ID
def extract_playlist_id(text):
    if not text:
        return None
    text = text.strip()
    if re.fullmatch(r'\d+', text):
        return text
    m = re.search(r'(\d{5,})', text)
    if m:
        return m.group(1)
    return None

# --- 与 QQ音乐接口交互（返回 (playlist_title, [(name,singers,album), ...], author) 或 None） ---

# 老接口: c.y.qq.com 获取歌单信息 返回 (title, song_list, author) 或 None
def try_c_y_qq(disstid):
    url = "https://c.y.qq.com/qzone/fcg-bin/fcg_ucc_getcdinfo_byids_cp.fcg"
    params = {
        "disstid": str(disstid),
        "type": "1",
        "json": "1",
        "utf8": "1",
        "onlysong": "0",
        "format": "json"
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"https://y.qq.com/n/ryqq/playlist/{disstid}"
    }

    try:
        resp = requests.get(url, params=params, headers=headers, timeout=10)
    except Exception:
        return None

    text = strip_jsonp(resp.text.strip())
    try:
        j = json.loads(text)
    except Exception:
        return None

    title = None
    author = ""
    if isinstance(j, dict):
        if "cdlist" in j and isinstance(j["cdlist"], list) and j["cdlist"]:
            cd = j["cdlist"][0]
            title = cd.get("dissname") or cd.get("diss_name") or cd.get("title") or cd.get("cdtitle") or cd.get("name")
            author = cd.get("data_signer") or cd.get("nickname") or cd.get("nick") or cd.get("username") or cd.get("uname") or ""
            songlist = cd.get("songlist") or cd.get("list") or []
        else:
            title = j.get("dissname") or j.get("title") or j.get("name")
            author = j.get("data_signer") or j.get("nickname") or j.get("nick") or ""
            songlist = j.get("songlist") or []
    else:
        return None

    results = []
    for s in songlist:
        name = s.get("songname") or s.get("name") or s.get("title") or ""
        if "singer" in s and isinstance(s["singer"], list):
            singers = ", ".join([a.get("name") or a.get("singer_name") or a.get("nickname","") for a in s["singer"]])
        else:
            singers = s.get("singername") or s.get("singer_name") or s.get("lan") or s.get("singer") or ""
        album = s.get("albumname") or (s.get("album") or {}).get("name") or s.get("albumname_utf8") or ""
        results.append((name, singers, album))

    return (title or "", results, author)

# 新接口: u.y.qq.com 的 GetPlaylistDetail 返回 (title, song_list, author) 或 None
def try_u_y_qq_playlist_detail(playlist_id):
    url = "https://u.y.qq.com/cgi-bin/musicu.fcg"
    payload = {
        "comm": {
            "cv": 4747474,
            "ct": 24,
            "format": "json",
            "inCharset": "utf-8",
            "outCharset": "utf-8",
            "notice": 0,
            "platform": "yqq.json",
            "needNewCode": 1,
            "uin": "0"
        },
        "playlist": {
            "method": "GetPlaylistDetail",
            "module": "music.playlist.PlaylistDetailServer",
            "param": {
                "id": int(playlist_id),
                "n": 1000,
                "order": 5
            }
        }
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"https://y.qq.com/n/ryqq/playlist/{playlist_id}",
        "Content-Type": "application/json"
    }
    try:
        resp = requests.post(url, data=json.dumps(payload), headers=headers, timeout=10)
    except Exception:
        return None

    try:
        j = resp.json()
    except Exception:
        return None

    pl = j.get("playlist")
    if not pl:
        return None

    title = ""
    author = ""
    if isinstance(pl, dict):
        title = pl.get("title") or pl.get("dissname") or (pl.get("data") or {}).get("title") or (pl.get("data") or {}).get("name") or pl.get("name") or ""
        creator = pl.get("creator") or (pl.get("data") or {}).get("creator") or {}
        if isinstance(creator, dict):

            # 兼容各种字段名
            author = creator.get("nickname") or creator.get("nickName") or creator.get("nick") or creator.get("name") or creator.get("data_signer") or ""
        if not author:
            author = pl.get("data_signer") or pl.get("nickname") or pl.get("nick") or ""

    if pl.get("code") is not None and pl.get("code") != 0:
        return None

    songlist = (pl.get("data") or {}).get("songlist") or (pl.get("data") or {}).get("songs") or pl.get("songlist") or []
    results = []
    for s in songlist:
        name = s.get("name") or s.get("songname") or ""
        if isinstance(s.get("singer"), list):
            singers = ", ".join([a.get("name","") for a in s["singer"]])
        else:
            singers = s.get("singername") or s.get("singer") or ""
        album = (s.get("album") or {}).get("name") or s.get("albumname") or ""
        results.append((name, singers, album))
    return (title or "", results, author)

def get_playlist_songs(playlist_id):

    # 依次尝试不同接口，返回 (title, songs, author) 或 None
    res = try_c_y_qq(playlist_id)
    if res:
        return res

    res = try_u_y_qq_playlist_detail(playlist_id)
    if res:
        return res

    # 备用 payload（playlist_songlist），尝试从 data 中提取 author
    url = "https://u.y.qq.com/cgi-bin/musicu.fcg"
    payload_alt = {
        "comm": {"cv": 4747474, "ct": 24, "format": "json", "uin": "0", "platform": "yqq.json"},
        "playlist_songlist": {
            "method": "GetPlaylistSongs",
            "module": "playlist.PlaylistSongListSrv",
            "param": {"id": int(playlist_id), "start": 0, "count": 1000}
        }
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"https://y.qq.com/n/ryqq/playlist/{playlist_id}",
        "Content-Type": "application/json"
    }
    try:
        resp = requests.post(url, data=json.dumps(payload_alt), headers=headers, timeout=10)
        j = resp.json()
        p = j.get("playlist_songlist", {})
        if p.get("code") == 0:
            data = p.get("data", {})
            songlist = data.get("songlist", []) or []
            title = data.get("title") or data.get("dissname") or data.get("name") or ""
            results = []
            for s in songlist:
                name = s.get("name","")
                singers = ", ".join([a.get("name","") for a in s.get("singer", [])]) if isinstance(s.get("singer"), list) else s.get("singer","")
                album = (s.get("album") or {}).get("name","")
                results.append((name, singers, album))
            author = data.get("data_signer") or data.get("nickname") or data.get("nick") or data.get("username") or ""
            return (title or "", results, author)
    except Exception:
        pass

    return None

# --- 导出函数（txt/csv/xlsx/json） ---

# 保存为 txt（每行 "Title - Artist - Album"）
def export_to_txt(rows, output_path):
    header = ""
    try:
        with open(output_path, "w", encoding="utf-8-sig", newline="\n") as f:
            f.write(header)
            for name, singers, album in rows:
                safe_name = (name or "").replace("\r"," ").replace("\n"," ")
                safe_singers = (singers or "").replace("\r"," ").replace("\n"," ")
                safe_album = (album or "").replace("\r"," ").replace("\n"," ")
                f.write(f"{safe_name} - {safe_singers} - {safe_album}\n")
    except Exception:
        raise

# 保存为 csv（utf-8-sig，第一行为表头）
def export_to_csv(rows, output_path):
    try:
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Title", "Artist", "Album"])
            for name, singers, album in rows:
                safe_name = (name or "").replace("\r"," ").replace("\n"," ")
                safe_singers = (singers or "").replace("\r"," ").replace("\n"," ")
                safe_album = (album or "").replace("\r"," ").replace("\n"," ")
                writer.writerow([safe_name, safe_singers, safe_album])
    except Exception:
        raise

# 保存为 xlsx（openpyxl），第一行为表头
def export_to_xlsx(rows, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise

    wb = Workbook()
    ws = wb.active
    ws.title = "QQ音乐歌单"
    ws.append(["Title", "Artist", "Album"])
    for name, singers, album in rows:
        ws.append([name or "", singers or "", album or ""])
    for col_idx in range(1, 4):
        col = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col]:
            try:
                l = len(str(cell.value))
            except:
                l = 0
            if l > max_len:
                max_len = l
        ws.column_dimensions[col].width = min(max(10, int(max_len * 1.1) + 2), 60)
    wb.save(output_path)

# 保存为 json（数组对象），使用 utf-8 编码
def export_to_json(rows, output_path):
    data = []
    for name, singers, album in rows:
        data.append({
            "Title": name or "",
            "Artist": singers or "",
            "Album": album or ""
        })
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        raise

# 打开并选中文件（Windows），或在其他平台打开文件所在目录
def open_file_location(path):
    try:
        abs_path = os.path.abspath(path)
        dirpath = os.path.dirname(abs_path)
        system = platform.system()
        if system == "Windows":

            # explorer /select, 可以高亮选中文件
            try:
                subprocess.run(["explorer", "/select,", abs_path])
                return
            except Exception:

                # 退回到打开目录
                try:
                    os.startfile(dirpath)
                    return
                except Exception:
                    pass
        elif system == "Darwin":

            # macOS
            try:
                subprocess.run(["open", dirpath])
                return
            except Exception:
                pass
        else:

            # Linux / 其他，尝试 xdg-open
            try:
                subprocess.run(["xdg-open", dirpath])
                return
            except Exception:
                pass
    except Exception:
        pass

# --- 主程序入口（循环式：导出完一个文件后继续输入，输入 0 / q / quit / exit 退出） ---
def main():
    print("============ QQ音乐歌单导出工具 ============")
    try:
        while True:
            user_input = input("请输入歌单链接 或 歌单ID（输入 0 退出）：").strip()
            if user_input.lower() in ("0", "q", "quit", "exit"):
                print("\n========== 程序已退出，感谢使用！===========")
                break

            pid = extract_playlist_id(user_input)
            if not pid:
                print("无法从输入中提取歌单ID，请确认链接或直接输入数字ID\n")
                print(f"==========================================")
                continue

            res = get_playlist_songs(pid)
            if not res:
                print("\n歌单不存在或无法读取")
                print(f"==========================================")
                continue

            title, songs, author = res  # 解包 author
            if not songs:
                print("\n歌单不存在或无法读取")
                print(f"==========================================")
                continue

            playlist_title = title.strip() if title else f"playlist_{pid}"
            safe_title = sanitize_filename(playlist_title)

            # 处理 author 显示与用于文件名的安全化
            display_author = author or "未知"
            safe_author = sanitize_filename(display_author) or "未知作者"

            print(f"\n已获取到 {display_author} 的歌单：\n名称：{playlist_title}，共 {len(songs)} 首歌曲")
            print(f"==========================================")

            # 显示数字菜单供用户选择
            print("请选择导出格式：")
            print(" 1) .xlsx  - (默认) Excel 文件")
            print(" 2) .csv   - 标准 CSV utf-8-sig")
            print(" 3) .json  - JSON 文件，数组")
            print(" 4) .txt   - 纯文本格式")
            print(f"==========================================")
            choice = input("选择 (1 - 4，输入 0 退出程序)：").strip()

            if choice == "":
                choice = "1"
            if choice not in ("0","1","2","3","4"):
                print("选择无效，默认使用 xlsx")
                choice = "1"

            try:
                out_name = None

                if choice == "0":
                    print("\n========== 程序已退出，感谢使用！===========")
                    return

                elif choice == "1":
                    out_name = f"{safe_title} - {safe_author}.xlsx"  # [修改：文件名加入作者]
                    try:
                        export_to_xlsx(songs, out_name)
                        print(f"\n已保存为: {out_name}")
                        print(f"==========================================")

                    except ImportError:
                        print("导出 xlsx 失败：缺少 openpyxl 库，请运行：pip install openpyxl")
                        out_name = None

                elif choice == "2":
                    out_name = f"{safe_title} - {safe_author}.csv"  # [修改]
                    export_to_csv(songs, out_name)
                    print(f"\n已保存为: {out_name}")
                    print(f"==========================================")

                elif choice == "3":
                    out_name = f"{safe_title} - {safe_author}.json"  # [修改]
                    export_to_json(songs, out_name)
                    print(f"\n已保存为: {out_name}")
                    print(f"==========================================")

                elif choice == "4":
                    out_name = f"{safe_title} - {safe_author}.txt"  # [修改]
                    export_to_txt(songs, out_name)
                    print(f"\n已保存为: {out_name}")
                    print(f"==========================================")

                # 导出成功且有 out_name 时，尝试在系统中打开文件所在目录（Windows 会选中文件）
                if out_name:
                    open_file_location(out_name)

            except Exception as e:
                print("保存文件时出错：", e)

    except KeyboardInterrupt:
        print("\n\n========= 程序被用户终止，感谢使用！=========")

if __name__ == "__main__":
    main()